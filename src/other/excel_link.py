# -*- coding: utf-8 -*-
from openpyxl import Workbook


def write_excel(row, col, value, link, ws):
    # 在A1单元格插入超链接
    ws.cell(row=row, column=col).hyperlink = link
    ws.cell(row=row, column=col).value = value


if __name__ == "__main__":
    wav_path = "to/2023-07-07_15533442510_E_拒绝次数 _= 0次.wav"
    asr_path = "to/asr.txt"

    wb = Workbook()
    ws = wb.active

    write_excel(1, 1, "record_test.wav", wav_path, ws)
    write_excel(1, 2, "asr_text.txt", asr_path, ws)

    wb.save("example1.xlsx")
